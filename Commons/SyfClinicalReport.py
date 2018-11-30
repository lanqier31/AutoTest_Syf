# -*- coding: utf-8 -*-
# Author：WeirGao

from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.select import Select
from selenium.webdriver.support import expected_conditions as EC
from openpyxl.reader.excel import load_workbook
from Config import reportType
from collections import OrderedDict
from datetime import datetime
# from openpyxl.styles import Font, colors, Alignment
import Config
import time,os
import operateExcel,globals
from time import sleep
#进入初诊页面

driver=Config.ChromeDriver
autocase = Config.autocase_path
book = load_workbook(autocase)

#病历号输入框
txtHid ="$('#txtHospitalNumber')"

def goto_Report():

    url = 'http://'+Config.IP+Config.Version+'/SyfHospitalClinicalDataCenter/QueryReport?type=add&editmark=true'
    driver.get(url)

def input_Hid(hid):
    '''输入病历号'''
    # exec_js = "$('#txtHospitalNumber').val(hid)"
    driver.execute_script(txtHid+'.val("' + hid + '")')
    alert=alert_close()
    if alert =='不存在该病历号':
        globals.log(str(hid)+u'没有此病人信息')
        return u'没有此病人信息'
    driver.execute_script(txtHid+".blur()")
    sleep(5)
    # WebDriverWait(driver, 10).until(
    #     lambda the_driver: the_driver.find_element_by_id('selShouShuList').is_displayed())
    alert_close()

def num_surgery():
    '''检查手术次数'''
    surgeryList =Select(driver.find_element_by_id('selShouShuList'))
    options_list = surgeryList.options
    return len(options_list)


def alert_close():
    """判断是否存在提醒框并关闭"""
    alert = EC.alert_is_present()(driver)
    if alert:
        if (u'外院手术' in alert.text):
            alert.dismiss()
        elif (u'没有此病人信息'in alert.text):
            alert.accept()
            return "不存在该病历号"
        elif u'没有抓取到报告内容'in alert.text:
            alert.accept()

        else:
            alert.accept()

#截图
def screenPage(reportType):
    """type:包含config 中 reporttype的key"""
    nowTime = time.strftime("%Y%m%d.%H.%M.%S")
    dirs = r'../PageScreen/' + reportType
    if not os.path.exists(dirs):
        os.makedirs(dirs)
    driver.get_screenshot_as_file(dirs + '/%s.png' % nowTime)


def undo():
    driver.find_element_by_id('btnRes').click()
    sleep(1)
    alert_close()
    sleep(1)


def selectType(type):
    """ 报告类型列表中选择报告类型"""
    wait_loading()
    driver.find_element_by_id('txtCheckReportTypeA').click()
    sleep(1)
    clickNum = 0
    reporttype = driver.find_element_by_css_selector('div[data-text="' +reportType[type]+ '"]')
    while not (reporttype.is_displayed()):
        if clickNum ==2:
            return "该报告类别没有找到"
        else:
            driver.find_element_by_id('txtCheckReportTypeA').click()
            clickNum= clickNum+1
    reporttype.click()  # 选择该报告类别
    WebDriverWait(driver, 10).until(
        lambda the_driver: the_driver.find_element_by_id('tbodyReportList').is_displayed())
    sleep(1)


def yearSelect(shiduan,value):
    """shiduan：a:年期内；b：测评段；c：孕药控；
        value：2；5；10；"""
    sleep(2)
    alert_close()
    WebDriverWait(driver, 10).until(
        lambda the_driver: the_driver.find_element_by_id('divOperationList').is_displayed())
    # 读取首次手术下的随访
    try:
        table = driver.find_element_by_xpath('//div[@id="divOperationList"]/table[last()]')
        sleep(1)
        table.find_element_by_xpath('tbody/tr[4]/td[1]').click()
        sleep(1)
        #
        # table.find_element_by_name('selNianQi')
        selshiduan = Select(table.find_element_by_name("selShiDuan"))
        selshiduan.select_by_value(shiduan)
        year = Select(table.find_element_by_name('selNianQi'))
        year.select_by_value(value)
        sleep(1)
        alert_close()
    except Exception as e:
        print e


def jiaoyan_Bchao(Hid):
    sheet = book['Bchao']
    n = operateExcel.max_row('Bchao') + 1
    try:
        WebDriverWait(driver, 10).until(
            lambda the_driver: the_driver.find_element_by_id('tbodyReportList').is_displayed())
    except Exception as e:
        print e
    alert_close()
    tbodyReportList = driver.find_element_by_id('tbodyReportList')
    reportList = tbodyReportList.find_elements_by_tag_name('tr')
    if(len(reportList)== 1):
        if (reportList[0].text=='没有报告信息.'):
            return Hid + u'无B超报告'
    for j in range(len(reportList)):  # 遍历B超报告份数
        wait_loading()
        reportList = tbodyReportList.find_elements_by_tag_name('tr')
        if (len(reportList) <= j):
            return (Hid + u"报告日期超出范围")
        sleep(3)
        reportList[j].click()
        sleep(1)
        # 判断是否有alert
        alert = EC.alert_is_present()(driver)
        if alert:
            if (u'外院手术' in alert.text):
                alert.dismiss()
            else:
                alert.accept()
        className = reportList[j].find_element_by_xpath('td[3]/div').get_attribute("class")
        while ('StateNoCom' != className and 'StateCFCom' != className):
            wait_loading()
            undo()
            alert_close()
            wait_loading()
            reportList = tbodyReportList.find_elements_by_tag_name('tr')
            if(len(reportList)<=j):
                return (Hid+u"报告日期超出范围")
            reportList[j].click()  # 焦点重新回到该报告上
            alert_close()
            className = reportList[j].find_element_by_xpath('td[3]/div').get_attribute("class")
        wait_loading()
        checkResult = driver.find_element_by_xpath('//div[@id="divUltrasonography"]/div[1]/div[5]/div[2]').text
        checkConclusion = driver.find_element_by_xpath(
            '//div[@id="divUltrasonography"]/div[1]/div[6]/div[2]').text
        WebDriverWait(driver, 10).until(
            lambda the_driver: the_driver.find_element_by_id('btnTest').is_displayed())
        jiaoyan_format_Bchao()  #B超校验时对淋巴结进行格式化处理
        result = readBCtext()
        sheet['A'+str(n)] = Hid
        sheet['B'+str(n)] = checkResult
        sheet['C'+str(n)] = checkConclusion
        sheet['D'+str(n)] = result['suoj_xy']
        sheet['E'+str(n)] = result['zhend_xy']
        sheet['F'+str(n)] = result['suoj_xc']
        sheet['G'+str(n)] = result['zhend_xc']
        sheet['H'+str(n)] = result['suoj_qcI']
        sheet['I'+str(n)] = result['zhend_qcI']
        sheet['J' + str(n)] = result['suoj_qcI']
        sheet['K' + str(n)] = result['zhend_qcI']
        sheet['L'+str(n)] = result['suoj_pl']
        sheet['M'+str(n)] = result['zhend_pl']
        book.save(autocase)
        n = n+1
        screenPage('Bchao')
        driver.find_element_by_id('btnCode').click()  # 点击代码化
        sleep(3)
        Codepai_Bchao() # 代码化中所有的派生
        screenPage('Bchao')
        driver.find_element_by_id('btnQuery').click() #返回
        WebDriverWait(driver, 10).until(
            lambda the_driver: the_driver.find_element_by_id('tbodyReportList').is_displayed())


def jiaoyan(reportType,Hid):
    """通用类型：Img131I，ImgA，ImgB"""
    sheet = book[reportType]
    row = operateExcel.max_row(reportType) + 1
    try:
        WebDriverWait(driver, 10).until(
            lambda the_driver: the_driver.find_element_by_id('tbodyReportList').is_displayed())
    except Exception as e:
        print e
    alert_close()
    tbodyReportList = driver.find_element_by_id('tbodyReportList')
    reportList = tbodyReportList.find_elements_by_tag_name('tr')
    if(len(reportList)== 1):
        if (reportList[0].text=='没有报告信息.'):
            return Hid + u'无报告'
    for j in range(len(reportList)):  # 遍历报告份数
        wait_loading()
        reportList = tbodyReportList.find_elements_by_tag_name('tr')
        if (len(reportList) <= j):
            return (Hid + u"报告日期超出范围")
        sleep(3)
        reportList[j].click()
        sleep(1)
        # 判断是否有alert
        alert = EC.alert_is_present()(driver)
        if alert:
            if (u'外院手术' in alert.text):
                alert.dismiss()
            else:
                alert.accept()
        className = reportList[j].find_element_by_xpath('td[3]/div').get_attribute("class")
        while ('StateNoCom' != className and 'StateCFCom' != className):
            wait_loading()
            undo()
            alert_close()
            wait_loading()
            reportList = tbodyReportList.find_elements_by_tag_name('tr')
            if(len(reportList)<=j):
                return (Hid+u"报告日期超出范围")
            reportList[j].click()  # 焦点重新回到该报告上
            alert_close()
            className = reportList[j].find_element_by_xpath('td[3]/div').get_attribute("class")
        wait_loading()
        yuanshiText = readyuanshiReport(reportType)
        jiaoyanText = readjiaoyanReport(reportType)
        zimu = ['B','C','D','E','F','G','H','I','J','K']
        sheet['A'+str(row)]=Hid
        i=0
        for t in yuanshiText:
            sheet[zimu[i]+str(row)]=yuanshiText[t]
            i=i+1
        for s in jiaoyanText:
            sheet[zimu[i]+str(row)]=jiaoyanText[s]
            i=i+1
        row=row+1
        book.save(autocase)
        screenPage(reportType)


def readBC_pretext():
    """"读取超声校验报告中各个框的值"""
    suoj_xy = "return $('div[name=" 'checkResult' "]').html();"
    zhend_xy = "return $('div[name=" 'checkConclusion' "]').html();"
    suoj_ln = "return $('div[name=" 'checkResult2' "]').html();"
    zhend_ln = "return $('div[name=" 'checkConclusion2' "]').html();"

    return {
        "suoj_xy": driver.execute_script(suoj_xy),
        "zhend_xy": driver.execute_script(zhend_xy),
        "suoj_ln": driver.execute_script(suoj_ln),
        "zhend_ln": driver.execute_script(zhend_ln),
    }

def readBC_fellowuptext():
    """"读取超声校验报告中各个框的值"""
    suoj_xy = "return $('div[name=" 'checkResult' "]').html();"
    zhend_xy = "return $('div[name=" 'checkConclusion' "]').html();"
    suoj_xc = "return $('div[name=" 'checkResultA' "]').html();"
    zhend_xc = "return $('div[name=" 'checkConclusionA' "]').html();"
    suoj_qcI = "return $('div[name=" 'checkResultBlockI' "]').html();"
    zhend_qcI = "return $('div[name=" 'checkConclusionBlockI' "]').html();"
    suoj_qcII = "return $('div[name=" 'checkResult2' "]').html();"
    zhend_qcII = "return $('div[name=" 'checkConclusion2' "]').html();"
    suoj_pl = "return $('div[name=" 'checkResult2A' "]').html();"
    zhend_pl = "return $('div[name=" 'checkConclusion2A' "]').html();"
    return {
        "suoj_xy":driver.execute_script(suoj_xy),
        "zhend_xy":driver.execute_script(zhend_xy),
        "suoj_xc":driver.execute_script(suoj_xc),
        "zhend_xc":driver.execute_script(zhend_xc),
        "suoj_qcI":driver.execute_script(suoj_qcI),
        "zhend_qcI":driver.execute_script(zhend_qcI),
        "suoj_qcII": driver.execute_script(suoj_qcII),
        "zhend_qcII": driver.execute_script(zhend_qcII),
        "suoj_pl":driver.execute_script(suoj_pl),
        "zhend_pl":driver.execute_script(zhend_pl),
    }


def readImg131Itext():
    """"读取核素影像中校验各个分框的值"""
    fo = "return $('div[name=" 'checkResult' "]').html();"
    neckAFT = "return $('div[name=" 'NeckAFT' "]').html();"
    lung = "return $('div[name=" 'checkBothLungMediastinum' "]').html();"
    quanshen = "return $('div[name=" 'checkWholeSkeleton' "]').html();"
    Skeleton = "return $('div[name=" 'checkWholebody' "]').html();"
    ct = "return $('div[name=" 'checkMachineCT' "]').html();"
    result = "return $('div[name=" 'checkConclusion' "]').html();"

    return {
        "fo":driver.execute_script(fo),
        "neckAFT":driver.execute_script(neckAFT),
        "lung":driver.execute_script(lung),
        "quanshen":driver.execute_script(quanshen),
        "Skeleton":driver.execute_script(Skeleton),
        "ct":driver.execute_script(ct),
        "result": driver.execute_script(result),
    }


def ImgAjiaoyantext():
    """读取影像A校验报告中的各个所见框的值"""
    fo = "return $('div[name=" 'checkResult' "]').html();"
    AFT = "return $('div[name=" 'AFT_YxA' "]').html();"
    larbloodVessels = "return $('div[name=" 'larbloodVessels' "]').html();"
    TrachEsophagus = "return $('div[name=" 'TrachEsophagus' "]').html();"
    return {
        "fo":driver.execute_script(fo),
        "AFT":driver.execute_script(AFT),
        "larbloodVessels":driver.execute_script(larbloodVessels),
        "TrachEsophagus":driver.execute_script(TrachEsophagus),
    }


def ImgBjiaoyantext():
    """读取影像B校验报告中的各个所见框的值"""
    res = OrderedDict()

    suoj_fsz ="return $('div[name=" 'checkResult' "]').html();"
    suoj_fm = "return $('div[name=" 'checkResultFm' "]').html();"
    suoj_xxg = "return $('div[name=" 'checkResultXxg' "]').html();"
    suoj_qt = "return $('div[name=" 'checkResultFj' "]').html();"
    jielun = "return $('div[name=" 'checkConclusion' "]').html();"
    res['suoj_fsz'] = driver.execute_script(suoj_fsz)
    res['suoj_fm'] = driver.execute_script(suoj_fm)
    res['suoj_xxg'] = driver.execute_script(suoj_xxg)
    res['suoj_qt'] = driver.execute_script(suoj_qt)
    res['jielun'] = driver.execute_script(jielun)
    return res



def Codepai_Bchao():
    """"B超代码化中的派生"""
    codePai = {
    "xnz" : driver.find_element_by_xpath('//div[@id="divUltrasonography"]/div[5]/div[3]/div[1]/div'),
    "ln" : driver.find_element_by_xpath('//div[@id="divUltrasonography"]/div[5]/div[5]/div[1]/div'),
    "ln_qs" : driver.find_element_by_xpath('//div[@id="divUltrasonography"]/div[5]/div[6]/div[1]/div'),
    "ln_pl" : driver.find_element_by_xpath('//div[@id="divUltrasonography"]/div[5]/div[7]/div[1]/div'),
    "total" : driver.find_element_by_xpath('//div[@id="divUltrasonography"]/div[5]/div[8]/div[1]/div')}
    for key in codePai:
        if codePai[key].is_displayed():
            codePai[key].click()
            sleep(1)


def jiaoyan_format_Bchao():
    """B超校验时对淋巴结进行格式化处理"""
    order_ln1 = driver.find_element_by_xpath('//span[@data-cmd="InsertOrderedList-Bchao_01"]')
    order_ln2 = driver.find_element_by_xpath('//span[@data-cmd="InsertOrderedList-Bchao_02"]')
    if order_ln1.is_displayed():
        order_ln1.click()
        sleep(1)
    if order_ln2.is_displayed():
        order_ln2.click()
        sleep(1)


def is_element_visible(element):
    """判断是元素是否可见"""
    try:
        the_element = EC.visibility_of_element_located(element)
        assert the_element(driver)
        flag = True
    except:
        flag = False
    return flag


def wait_loading():
    """waiting for divBlockHid disappear"""
    WebDriverWait(driver, 30).until_not(
        lambda the_driver: the_driver.find_element_by_class_name('divBlockHid').is_displayed())


def readyuanshiReport(reportType):
    checkResult=''
    checkConclusion=''
    if reportType in ['Bchao_pre','Bchao_fellowup']:
        checkResult = driver.find_element_by_xpath('//div[@id="divUltrasonography"]/div[1]/div[5]/div[2]').text
        checkConclusion = driver.find_element_by_xpath(
            '//div[@id="divUltrasonography"]/div[1]/div[6]/div[2]').text
    if reportType =='ImgB':
        checkResult = driver.find_element_by_xpath('//div[@id="divImagingExamination"]/div[1]/div[5]/div[2]').text
        checkConclusion = driver.find_element_by_xpath(
            '//div[@id="divImagingExamination"]/div[1]/div[6]/div[2]').text
    if reportType=='ImgA':
        checkResult = driver.find_element_by_xpath('//div[@id="divImagingExaminationB"]/div[1]/div[5]/div[2]').text
        checkConclusion = driver.find_element_by_xpath(
            '//div[@id="divImagingExaminationB"]/div[1]/div[6]/div[2]').text
    if reportType =='Img131I':
        checkResult = driver.find_element_by_xpath('//div[@id="divNuclideImgExamination"]/div[1]/div[5]/div[2]').text
        checkConclusion = driver.find_element_by_xpath(
            '//div[@id="divNuclideImgExamination"]/div[1]/div[6]/div[2]').text
    if reportType == 'Bfna_cell':
        checkResult = driver.find_element_by_xpath('//div[@id="divImagingExamination"]/div[1]/div[5]/div[2]').text
        checkConclusion = driver.find_element_by_xpath(
            '//div[@id="divImagingExamination"]/div[1]/div[6]/div[2]').text

    return {
        "checkResult":checkResult,
        "checkConclusion":checkConclusion,
    }


def readjiaoyanReport(reportType):
    result={}
    if reportType =='Bchao_fellowup':
        result= readBC_fellowuptext()
    if reportType == 'Bchao_pre':
        result = readBC_pretext()
    if reportType == 'ImgA':
        result = ImgAjiaoyantext()
    if reportType == 'ImgB':
        result = ImgBjiaoyantext()
    if reportType == 'Img131I':
        result = readImg131Itext()

    return result





