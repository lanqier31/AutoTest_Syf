# -*-coding:utf-8-*-
import Config
from Commons import operateExcel
from openpyxl.reader.excel import load_workbook

autocase = Config.autocase_path
def queryreport():
    sa=[]
    book = load_workbook(autocase)
    sheet = book.get_sheet_by_name('ReportNo')
    if not sheet:
        book.create_sheet('ReportNo', index=1)
    conn = Config.conn
    cursor = conn.cursor()
    sql = "SELECT HospitalNumber,ReportNo FROM [HospitalOriginalReportList] where ReportName like '%常规组织%' and State != 99 and ORDERNO != 'waiyuan'"
    cursor.execute(sql)
    row = cursor.fetchall()
    n=2
    for i in row:
        sheet['A'+str(n)] = i[0]
        sheet['B'+str(n)] = i[1]
        n = n+1
    # 关闭连接
    book.save(autocase)
    conn.close()

