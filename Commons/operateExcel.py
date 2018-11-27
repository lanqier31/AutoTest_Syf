#-*- Coding=utf-8 -*-
from openpyxl.reader.excel import load_workbook
from openpyxl import workbook
from Commons import globals
import Config
import os
import time

import sys
reload(sys)
sys.setdefaultencoding('utf-8')

autocase = Config.autocase_path
def WriteExcel(result, locator,sheetname):
    book = load_workbook(autocase)
    sheet = book.get_sheet_by_name(sheetname)
    sheet[locator]= result


def SaveExcel(excelName):
    book = load_workbook(excelName)
    book.save(excelName)


def ReadExcel(locator,sheetname):
    book = load_workbook(autocase)
    sheet = book[sheetname]
    content = sheet[locator].value
    return content


def max_row(sheetname):
    book = load_workbook(autocase)
    sheet = book[sheetname]
    return sheet.max_row


def max_column(sheetname):
    book = load_workbook(autocase)
    sheet = book[sheetname]
    return sheet.max_column


def All_content(sheetname):
    contents=[]
    book = load_workbook(autocase)
    sheet = book[sheetname]
    for row in sheet.rows:
        for cell in row:
            con = str(cell.value)
            contents.append(con)
    return contents[1:]

def get_column(sheetname):
    contents = []
    book = load_workbook(autocase)
    sheet = book[sheetname]
    cols = list(sheet.columns)[1]
    for cell in cols:
        con = str(cell.value)
        contents.append(con)
    return contents