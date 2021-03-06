#!/usr/bin/env python
# -*-coding:utf-8 -*-
# Author:  Weir Gao --<>
# Purpose:
# Created: 2018/7/24


from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import time
from time import sleep
from openpyxl.reader.excel import load_workbook
import Config
from Commons import Login, SyfClinicalReport,ReportList,operateExcel,globals


#we use global varialbe driver
driver=Config.ChromeDriver
#Url for login is a global varible
url=Config.LoginUrl
autocase = Config.autocase_path
log = Config.log_file_path
book = load_workbook(autocase)
sheet = book['pathology']
Login.login_Syf(url,'30048','5913')
Login.maxmize_window()
HospitalNums = operateExcel.All_content('Hid')

n = operateExcel.max_row('pathology')+1  # excel row
SyfClinicalReport.goto_Report()
for Hid in HospitalNums:   #遍历要测试的病历号
    # ReportList.goto_reportList()
    # ReportList.del_checkCode(Hid)        #删除该病历号下的校验代码化内容
    SyfClinicalReport.input_Hid(Hid)
    is_disappeared = WebDriverWait(driver, 20, 1).until_not(
        lambda x: x.find_element_by_xpath('//div[@class="divBlockHid"]').is_displayed())
    if is_disappeared:
        globals.log(Hid+' '+"手术信息加载超时")
        continue
    surgeryList = driver.find_element_by_id('selShouShuList')
    num_operations= SyfClinicalReport.num_surgery()
    if num_operations == 0:
        globals.log(str(num_operations))
        globals.log(Hid+"手术次数未获取到")
        continue
    for i in range(num_operations):  # 遍历手术次数
        SyfClinicalReport.wait_loading()
        surgeryList.click()
        # 判断是否有alert并关闭
        SyfClinicalReport.alert_close()
        surgeryList.find_elements_by_tag_name('option')[i].click()
        sleep(2)
        #判断是否有alert并关闭
        SyfClinicalReport.alert_close()
        SyfClinicalReport.wait_loading()
        driver.find_element_by_id('txtCheckReportType').click()
        driver.find_element_by_css_selector("div[data-text='常规病理']").click()  # 选择病理形态学的报告类别
        sleep(3)
        SyfClinicalReport.alert_close()
        tbodyReportList = driver.find_element_by_id('tbodyReportList')
        reportList = tbodyReportList.find_elements_by_tag_name('tr')
        if (len(reportList) == 1):
            if (reportList[0].text == '没有报告信息.'):
                globals.log(str(Hid) + u'无报告内容')
                continue
        for j in range(len(reportList)):    #遍历报告份数
            try:
                reportList = tbodyReportList.find_elements_by_tag_name('tr')
                SyfClinicalReport.wait_loading()
                # ActionChains(driver).move_to_element(reportList[j]).perform()
                reportList[j].click()
                sleep(2)
                SyfClinicalReport.alert_close()
                SyfClinicalReport.wait_loading()
                reportTitle = reportList[j].find_elements_by_tag_name('td')[0].get_attribute('title')
                if (u'冰冻' in reportTitle):
                    className = reportList[j].find_element_by_xpath('td[3]/div').get_attribute("class")
                    while ('StateNoCom' != className and 'StateCFCom' != className):
                        SyfClinicalReport.wait_loading()
                        SyfClinicalReport.undo()
                        sleep(2)
                        SyfClinicalReport.alert_close()
                        reportList = tbodyReportList.find_elements_by_tag_name('tr')
                        SyfClinicalReport.wait_loading()
                        reportList[j].click()  # 焦点重新回到该报告上
                        className = reportList[j].find_element_by_xpath('td[3]/div').get_attribute("class")
                    checkResult = driver.find_element_by_xpath('//div[ @ id = "divPathology"]/div[1]/div[5]/div[2]').text
                    checkConclusion = driver.find_element_by_xpath(
                        '//div[ @ id = "divPathology"]/div[1]/div[6]/div[2]').text
                    SyfClinicalReport.wait_loading()
                    # driver.find_element_by_id('btnTest').click()  # 点击校验按钮
                    suoj_Fo = driver.find_element_by_xpath('//div[@id="divPathology"]/div[3]/div[5]/div[1]/div[2]').text  # 腺内灶所见
                    zhend_Fo = driver.find_element_by_xpath('//div[@id="divPathology"]/div[3]/div[5]/div[2]/div[2]').text  # 腺内灶诊断
                    suoj_Ln = driver.find_element_by_xpath('//div[@id="divPathology"]/div[3]/div[5]/div[3]/div[2]').text  # 淋巴结所见
                    zhend_Ln = driver.find_element_by_xpath('//div[@id="divPathology"]/div[3]/div[5]/div[4]/div[2]').text  # 淋巴结诊断

                    sheet['A'+str(n)] = Hid
                    sheet['B'+str(n)] = checkResult
                    sheet['C'+str(n)] = checkConclusion
                    sheet['D'+str(n)] = suoj_Fo
                    sheet['E'+str(n)] = zhend_Fo
                    sheet['G'+str(n)] = suoj_Ln
                    sheet['I'+str(n)] = zhend_Ln

                    SyfClinicalReport.screenpatho()
                    driver.find_element_by_xpath('//span[@data-cmd="InsertOrderedList-xnz"]').click()
                    driver.find_element_by_xpath('//span[@data-cmd="InsertOrderedList-ln"]').click()
                    sleep(1)
                    sort_Fo = driver.find_element_by_xpath('//div[@id="divPathology"]/div[3]/div[5]/div[2]/div[2]').text   #排序后腺内灶
                    sort_Ln = driver.find_element_by_xpath('//div[@id="divPathology"]/div[3]/div[5]/div[3]/div[2]').text    #排序后淋巴结

                    sheet['F'+str(n)] = sort_Fo
                    sheet['H'+str(n)] = sort_Ln
                    n = n + 1
                    # SyfClinicalReport.screenpatho()
                    book.save(autocase)
                    driver.find_element_by_id('btnCode').click()   #点击代码化
                    sleep(2)
                    alert = EC.alert_is_present()(driver)
                    if alert:
                        if u'请先合并'in alert.text:
                            alert.accept()
                            driver.find_element_by_id('btnSave').click()
                            sleep(2)
                            merge = driver.find_element_by_id('divSelReport')
                            if merge.is_displayed():
                                driver.find_element_by_id('checkAll').click()
                                driver.find_element_by_xpath('//div[@id="divSelReport"]/div/div[3]/span').click()
                                sleep(2)
                                SyfClinicalReport.screenpatho()
                                break
                            else:
                                SyfClinicalReport.alert_close()
                        elif u'有未校验' in alert.text:
                            alert.accept()
                            driver.find_element_by_id('btnSave').click()
                            sleep(1)
                            SyfClinicalReport.alert_close()
                        else:
                            globals.log(Hid+alert.text)
                            alert.accept()
                    else:
                        sleep(2)
                        SyfClinicalReport.screenpatho()

                    # driver.find_element_by_id('btnQuery').click()
                    # WebDriverWait(driver, 20).until(
                    #     lambda the_driver: the_driver.find_element_by_id('tbodyReportList').is_displayed())
            except Exception,e:
                globals.log(str(Hid)+e.message)
