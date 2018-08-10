#!/usr/bin/env python
# -*-coding:utf-8 -*-
# Author:  Weir Gao --<>
# Purpose:
# Created: 2018/7/24

from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoAlertPresentException
import time
from time import sleep
import unittest
import sys
import os
import openpyxl
from openpyxl.reader.excel import load_workbook
import HTMLTestRunner

import Config
from Commons import Login, SyfClinicalReport,ReportList,operateExcel


#we use global varialbe driver
driver=Config.ChromeDriver
#Url for login is a global varible
url=Config.LoginUrl
autocase = Config.autocase_path
log = Config.log_file_path
book = load_workbook(autocase)
sheet = book['pathology']
Login.login_Syf(url,'30048','5913')
Login.maxmize_window()
HospitalNums = operateExcel.All_content('Hid')

n = operateExcel.max_row('pathology')+1  # excel row
for Hid in HospitalNums:   #遍历要测试的病历号
    # ReportList.goto_reportList()
    # ReportList.del_checkCode(Hid)        #删除该病历号下的校验代码化内容
    SyfClinicalReport.goto_Report()
    SyfClinicalReport.input_Hid(Hid)
    # 判断病历号是否存在
    is_disappeared = WebDriverWait(driver, 20, 1).until_not(
        lambda x: x.find_element_by_xpath('//div[@class="divBlockHid"]').is_displayed())
    if is_disappeared:
        log (Hid+' '+"手术信息加载超时")
        continue

    surgeryList = driver.find_element_by_id('selShouShuList')
    num_operations= SyfClinicalReport.num_surgery()
    if num_operations == 0:
        log(num_operations)
        log(Hid+"手术次数未获取到")
        continue
    for i in range(num_operations):  # 遍历手术次数
        WebDriverWait(driver, 20).until_not(
            lambda the_driver: the_driver.find_element_by_class_name('divBlockHid').is_displayed())
        surgeryList.click()
        # 判断是否有alert并关闭
        SyfClinicalReport.alert_close()
        surgeryList.find_elements_by_tag_name('option')[i].click()
        sleep(2)
        #判断是否有alert并关闭
        SyfClinicalReport.alert_close()
        WebDriverWait(driver, 20).until_not(
            lambda the_driver: the_driver.find_element_by_class_name('divBlockHid').is_displayed())
        driver.find_element_by_id('txtCheckReportType').click()
        driver.find_element_by_css_selector("div[data-text='病理形态学']").click()  # 选择病理形态学的报告类别
        sleep(3)
        SyfClinicalReport.alert_close()
        tbodyReportList = driver.find_element_by_id('tbodyReportList')
        reportList = tbodyReportList.find_elements_by_tag_name('tr')
        if (len(reportList) == 1):
            if (reportList[0].text == '没有报告信息.'):
                log(Hid + u'无报告内容')
        for j in range(len(reportList)):    #遍历报告份数
            reportList = tbodyReportList.find_elements_by_tag_name('tr')
            reportTitle = reportList[j].find_elements_by_tag_name('td')[0].get_attribute('title')
            if (u'常规组织病理' in reportTitle or u'手术标本' in reportTitle):
                reportList[j].click()
                sleep(1)
                # 判断是否有alert
                alert = EC.alert_is_present()(driver)
                if alert:
                    alert.accept()
                    continue
                WebDriverWait(driver, 20).until_not(
                    lambda the_driver: the_driver.find_element_by_class_name('divBlockHid').is_displayed())
                className = reportList[j].find_element_by_xpath('td[3]/div').get_attribute("class")
                while ('StateNoCom' != className and 'StateCFCom' != className):
                    WebDriverWait(driver, 20).until_not(
                        lambda the_driver: the_driver.find_element_by_class_name('divBlockHid').is_displayed())
                    SyfClinicalReport.undo()
                    sleep(2)
                    SyfClinicalReport.alert_close()
                    reportList = tbodyReportList.find_elements_by_tag_name('tr')
                    WebDriverWait(driver, 20).until_not(
                        lambda the_driver: the_driver.find_element_by_class_name('divBlockHid').is_displayed())
                    reportList[j].click()  # 焦点重新回到该报告上
                    className = reportList[j].find_element_by_xpath('td[3]/div').get_attribute("class")
                checkResult = driver.find_element_by_xpath('//div[ @ id = "divPathology"]/div[1]/div[5]/div[2]').text
                checkConclusion = driver.find_element_by_xpath(
                    '//div[ @ id = "divPathology"]/div[1]/div[6]/div[2]').text
                WebDriverWait(driver, 20).until_not(
                    lambda the_driver: the_driver.find_element_by_class_name('divBlockHid').is_displayed())

                driver.find_element_by_id('btnTest').click()  # 点击校验按钮
                suoj_Fo = driver.find_element_by_xpath('//div[@id="divPathology"]/div[3]/div[5]/div[2]').text  # 腺内灶所见
                zhend_Fo = driver.find_element_by_xpath('//div[@id="divPathology"]/div[3]/div[6]/div[2]').text  # 腺内灶诊断
                suoj_Ln = driver.find_element_by_xpath('//div[@id="divPathology"]/div[3]/div[7]/div[2]').text  # 淋巴结所见
                zhend_Ln = driver.find_element_by_xpath('//*[@id="divPathology"]/div[3]/div[8]/div[2]').text  # 淋巴结诊断

                sheet['A'+str(n)] = Hid
                sheet['B'+str(n)] = checkResult
                sheet['C'+str(n)] = checkConclusion
                sheet['D'+str(n)] = suoj_Fo
                sheet['E'+str(n)] = zhend_Fo
                sheet['G'+str(n)] = suoj_Ln
                sheet['I'+str(n)] = zhend_Ln

                SyfClinicalReport.screenAsTime()
                driver.find_element_by_xpath('//span[@data-cmd="InsertOrderedList-xnz"]').click()
                driver.find_element_by_xpath('//span[@data-cmd="InsertOrderedList-ln"]').click()
                sleep(1)
                sort_Fo = driver.find_element_by_xpath('//div[@id="divPathology"]/div[3]/div[6]/div[2]').text   #排序后腺内灶
                sort_Ln = driver.find_element_by_xpath('//div[@id="divPathology"]/div[3]/div[7]/div[2]').text    #排序后淋巴结

                sheet['F'+str(n)] = sort_Fo
                sheet['H'+str(n)] = sort_Ln
                n = n + 1
                SyfClinicalReport.screenAsTime()
                driver.find_element_by_id('btnCode').click()
                sleep(2)
                alert = EC.alert_is_present()(driver)
                if alert:
                        log(Hid+alert.text)
                        alert.dismiss()
                SyfClinicalReport.screenAsTime()
                book.save(autocase)
                driver.find_element_by_id('btnQuery').click()
                WebDriverWait(driver, 20).until(
                    lambda the_driver: the_driver.find_element_by_id('tbodyReportList').is_displayed())
